import os
import requests
from openpyxl import load_workbook
from PIL import Image
from tkinter import filedialog, Tk

# Function to download and convert images
def download_and_convert_images(excel_file, save_folder):
    # Load the Excel file
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # Open a log file to record failed downloads
    log_file_path = os.path.join(save_folder, 'failed_images.log')
    with open(log_file_path, 'w') as log_file:
        log_file.write("Failed Image Downloads Log\n")
        log_file.write("==========================\n")

        # Iterate through the rows and process each image
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row is the header
            sku, url = row
            try:
                # Download the image
                response = requests.get(url)
                response.raise_for_status()

                # Save the GIF image temporarily
                gif_path = os.path.join(save_folder, f"{sku}.gif")
                with open(gif_path, 'wb') as gif_file:
                    gif_file.write(response.content)

                # Convert the GIF to JPG
                with Image.open(gif_path) as img:
                    jpg_path = os.path.join(save_folder, f"{sku}.jpg")
                    img.convert('RGB').save(jpg_path, 'JPEG')

                # Remove the temporary GIF file
                os.remove(gif_path)

                print(f"Processed SKU: {sku}")
            except Exception as e:
                log_file.write(f"SKU: {sku}, URL: {url}, Error: {e}\n")
                print(f"Error processing SKU: {sku}, URL: {url}. Error: {e}")

# Main function to run the script
def main():
    # Hide the Tkinter root window
    root = Tk()
    root.withdraw()

    # Ask user for the Excel file
    excel_file = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx")])
    if not excel_file:
        print("No Excel file selected.")
        return

    # Ask user for the folder to save images
    save_folder = filedialog.askdirectory(title="Select Folder to Save Images")
    if not save_folder:
        print("No folder selected.")
        return

    # Download and convert images
    download_and_convert_images(excel_file, save_folder)
    print("Image processing complete.")

if __name__ == "__main__":
    main()
